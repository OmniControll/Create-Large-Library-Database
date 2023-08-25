import yfinance as yf
import matplotlib.pyplot as plt
import openpyxl

# Stock ticker symbols to fetch data for
tickers = ["MSFT", "INTC", "BTC-EUR", "ETH-EUR", "AMD", "NVDA"]

# Create a new Excel file
workbook = openpyxl.Workbook()

for ticker in tickers:
    # Get the stock data for the current ticker
    stock_data = yf.Ticker(ticker).history(period="3y")

    # Select the "Close" column and resample the data to get all rows per month
    monthly_data = stock_data["Close"].resample("M").mean()

    # Round the values to 2 decimal places
    monthly_data = monthly_data.round(2)

    # Remove the time stamps
    monthly_data.index = monthly_data.index.strftime("%Y-%m")

    # Convert the Series object to a dataframe
    monthly_data = monthly_data.to_frame()

    # Rename
    monthly_data.index.name = ticker
    monthly_data.columns = ["Price"]

    # Create a new sheet in the workbook for the current ticker
    sheet = workbook.create_sheet(title=ticker)

# Write the data to the sheet
sheet.append(["Ticker", "Price"])

for index, row in monthly_data.iterrows():
    sheet.append([index, row["Price"]])

# Save the workbook
workbook.save("stockdata.xlsx")

# Plot the data and save it to a file
plt.plot(monthly_data)
plt.title(ticker)
plt.grid(True)
plt.savefig(f"{ticker}.png")

# Open the Excel file
wb = openpyxl.load_workbook("stockdata.xlsx")

ws = wb.create_sheet(title=ticker)

# Iterate over the rows in the dataframe
for index, row in monthly_data.iterrows():
    ws.append([index, row["Price"]])


average_price = monthly_data["Price"].mean()

ws.append([ticker, "Average Price", average_price])

# Import the LineChart class
from openpyxl.chart import LineChart
from openpyxl.chart import Reference

# Create a new line chart
chart = LineChart()

# Set the chart properties
chart.title = "Stock Price"
chart.y_axis.title = "Price"
chart.x_axis.title = "Month"

# Add the data to the chart
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(monthly_data))
chart.add_data(data, titles_from_data=True)

# Insert the chart into the worksheet
ws.add_chart(chart, "E1")

wb.save("stockdata.xlsx")

